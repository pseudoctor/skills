#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票扫描器 - Excel 报销表生成器
根据 invoices.json 生成标准的费用报销明细表
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：未安装 openpyxl 库")
    print("请运行：pip3 install openpyxl")
    sys.exit(1)


def number_to_chinese(num):
    """将数字转换为中文大写"""
    chinese_num = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_unit = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']

    if num == 0:
        return '零圆整'

    str_num = str(int(num))
    result = ''
    length = len(str_num)

    for i, digit in enumerate(str_num):
        n = int(digit)
        if n != 0:
            result += chinese_num[n] + chinese_unit[length - i - 1]
        elif length - i - 1 > 0 and str_num[i + 1] != '0':
            result += '零'

    # 处理小数部分
    decimal_part = int((num - int(num)) * 100)
    if decimal_part > 0:
        jiao = decimal_part // 10
        fen = decimal_part % 10
        if jiao > 0:
            result += chinese_num[jiao] + '角'
        if fen > 0:
            result += chinese_num[fen] + '分'
    else:
        result += '整'

    return result + '圆' if not result.endswith('圆') else result


def format_date_for_excel(date_str, check_in=None, check_out=None):
    """格式化日期为Excel显示格式"""
    if check_in and check_out:
        # 住宿类：显示日期范围
        try:
            ci = datetime.strptime(check_in, '%Y-%m-%d')
            co = datetime.strptime(check_out, '%Y-%m-%d')
            return f"{ci.year}/{ci.month}/{ci.day}-{co.year}/{co.month}/{co.day}"
        except:
            pass

    if date_str:
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            return f"{dt.month}/{dt.day}/{str(dt.year)[2:]}"  # M/D/YY
        except:
            return date_str

    return "未知日期"


def map_category_to_excel(category, invoice_type):
    """映射类别到Excel显示"""
    mapping = {
        "市内交通": "交通费",
        "长途交通": "交通费",
        "住宿": "住宿费",
    }
    return mapping.get(category, invoice_type)


def create_expense_report(json_file, output_file, company_name=None, applicant_name=None):
    """创建费用报销明细表"""

    # 读取JSON数据
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "费用报销明细表"

    # 设置列宽
    ws.column_dimensions['A'].width = 6   # 序号
    ws.column_dimensions['B'].width = 12  # 类别
    ws.column_dimensions['C'].width = 15  # 日期
    ws.column_dimensions['D'].width = 18  # 名称
    ws.column_dimensions['E'].width = 12  # 金额
    ws.column_dimensions['F'].width = 8   # 单据数
    ws.column_dimensions['G'].width = 12  # 发票类型
    ws.column_dimensions['H'].width = 20  # 发票号码
    ws.column_dimensions['I'].width = 25  # 行程
    ws.column_dimensions['J'].width = 30  # 备注

    # 样式定义
    title_font = Font(name='宋体', size=14, bold=True)
    subtitle_font = Font(name='宋体', size=12, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    normal_font = Font(name='宋体', size=10)

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    current_row = 1

    # 第1行：公司名称
    if company_name:
        ws.merge_cells(f'A{current_row}:J{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = company_name
        cell.font = title_font
        cell.alignment = center_alignment
        current_row += 1

    # 第2行：表格名称
    ws.merge_cells(f'A{current_row}:J{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "费用报销明细表"
    cell.font = subtitle_font
    cell.alignment = center_alignment
    current_row += 1

    # 第3行：日期
    ws.merge_cells(f'A{current_row}:J{current_row}')
    cell = ws[f'A{current_row}']
    scan_date = data.get('scan_date', datetime.now().strftime('%Y-%m-%d'))
    dt = datetime.strptime(scan_date, '%Y-%m-%d')
    cell.value = f"{dt.year}/{dt.month}/{dt.day}"
    cell.font = normal_font
    cell.alignment = center_alignment
    current_row += 1

    # 第4行：表头
    headers = ['序号', '类别', '日期', '名称', '金额(元)', '单据数', '发票类型', '发票号码', '行程', '备注']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    header_row = current_row
    current_row += 1

    # 数据行
    invoices = data.get('invoices', [])
    total_amount = 0.0

    # 按类别排序：市内交通 -> 长途交通 -> 住宿 -> 其他
    category_order = {'市内交通': 1, '长途交通': 2, '住宿': 3, '其他': 4}
    sorted_invoices = sorted(invoices, key=lambda x: (
        category_order.get(x.get('category', '其他'), 5),
        x.get('date', '') or ''
    ))

    actual_seq = 0  # 实际序号计数器（修复序号跳号问题）
    for invoice in sorted_invoices:
        # 跳过重复和过滤的发票
        if invoice.get('duplicate') or invoice.get('filtered'):
            continue

        actual_seq += 1  # 序号递增
        row = current_row

        # A列：序号
        cell = ws.cell(row=row, column=1)
        cell.value = actual_seq  # 使用实际序号，保证连续
        cell.alignment = center_alignment
        cell.border = thin_border

        # B列：类别
        category = map_category_to_excel(
            invoice.get('category', '其他'),
            invoice.get('type', '')
        )
        cell = ws.cell(row=row, column=2)
        cell.value = category
        cell.alignment = center_alignment
        cell.border = thin_border

        # C列：日期
        date_str = format_date_for_excel(
            invoice.get('date'),
            invoice.get('check_in'),
            invoice.get('check_out')
        )
        cell = ws.cell(row=row, column=3)
        cell.value = date_str
        cell.alignment = center_alignment
        cell.border = thin_border

        # D列：名称
        name = invoice.get('type', '') or invoice.get('description', '').split('*')[-1]
        cell = ws.cell(row=row, column=4)
        cell.value = name
        cell.alignment = left_alignment
        cell.border = thin_border

        # E列：金额
        amount = invoice.get('amount', 0.0)
        total_amount += amount
        cell = ws.cell(row=row, column=5)
        cell.value = amount
        cell.number_format = '#,##0.00'
        cell.alignment = right_alignment
        cell.border = thin_border

        # F列：单据数
        receipt_count = invoice.get('receipt_count', 1)
        cell = ws.cell(row=row, column=6)
        cell.value = receipt_count
        cell.alignment = center_alignment
        cell.border = thin_border

        # G列：发票类型
        invoice_type = invoice.get('invoice_type', '')
        cell = ws.cell(row=row, column=7)
        cell.value = invoice_type
        cell.alignment = center_alignment
        cell.border = thin_border

        # H列：发票号码
        invoice_number = invoice.get('invoice_number', '')
        if not invoice_number:
            invoice_number = '-'
        cell = ws.cell(row=row, column=8)
        cell.value = invoice_number
        cell.alignment = center_alignment
        cell.border = thin_border

        # I列：行程
        route = ''
        if invoice.get('route'):
            route = invoice['route']
        elif invoice.get('from') and invoice.get('to'):
            route = f"{invoice['from']}-{invoice['to']}"
        cell = ws.cell(row=row, column=9)
        cell.value = route
        cell.alignment = center_alignment
        cell.border = thin_border

        # J列：备注
        warnings = invoice.get('validation_warnings', [])
        description = invoice.get('description', '')
        confidence = invoice.get('confidence', 1.0)

        remarks = []

        # 添加低置信度警告
        if confidence < 0.7:
            remarks.append(f"⚠️ 识别置信度较低({confidence:.0%})，建议人工核对")

        if warnings:
            remarks.extend(warnings)
        if description and invoice.get('category') == '其他':
            remarks.append(description)
        cell = ws.cell(row=row, column=10)
        cell.value = '\n'.join(remarks) if remarks else ''
        cell.alignment = left_alignment
        cell.border = thin_border

        # 低置信度行整行标黄
        if confidence < 0.7:
            low_confidence_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
            for col in range(1, 11):  # A到J列
                ws.cell(row=row, column=col).fill = low_confidence_fill

        current_row += 1

    data_end_row = current_row - 1

    # 汇总行
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "报销金额合计"
    cell.font = Font(name='宋体', size=11, bold=True)
    cell.alignment = center_alignment
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=5)  # 第5列是金额列
    cell.value = total_amount
    cell.number_format = '#,##0.00'
    cell.font = Font(name='宋体', size=11, bold=True)
    cell.alignment = right_alignment
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=6)
    cell.value = ""
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=7)
    cell.value = "人民币（大写）"
    cell.font = Font(name='宋体', size=10)
    cell.alignment = center_alignment
    cell.border = thin_border

    ws.merge_cells(f'H{current_row}:J{current_row}')
    cell = ws[f'H{current_row}']
    cell.value = number_to_chinese(total_amount)
    cell.font = Font(name='宋体', size=11, bold=True)
    cell.alignment = center_alignment
    cell.border = thin_border

    current_row += 2

    # 签字栏
    signature_row = current_row
    ws.merge_cells(f'A{signature_row}:B{signature_row}')
    cell = ws[f'A{signature_row}']
    cell.value = f"报销人：{applicant_name if applicant_name else ''}"
    cell.alignment = left_alignment

    ws.merge_cells(f'C{signature_row}:D{signature_row}')
    cell = ws[f'C{signature_row}']
    cell.value = "部门经理："
    cell.alignment = left_alignment

    ws.merge_cells(f'E{signature_row}:G{signature_row}')
    cell = ws[f'E{signature_row}']
    cell.value = "财务："
    cell.alignment = left_alignment

    ws.merge_cells(f'H{signature_row}:J{signature_row}')
    cell = ws[f'H{signature_row}']
    cell.value = "总经理："
    cell.alignment = left_alignment

    current_row += 2

    # 备注栏
    ws.merge_cells(f'A{current_row}:J{current_row + 2}')
    cell = ws[f'A{current_row}']
    cell.value = "备注："
    cell.alignment = Alignment(horizontal='left', vertical='top')
    cell.border = thin_border

    # 保存文件
    wb.save(output_file)
    print(f"✅ Excel报销表已生成：{output_file}")
    print(f"   - 发票数量：{len([i for i in invoices if not i.get('duplicate') and not i.get('filtered')])}")
    print(f"   - 总金额：¥{total_amount:,.2f}")
    print(f"   - 大写金额：{number_to_chinese(total_amount)}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='生成费用报销明细表（Excel格式）')
    parser.add_argument('json_file', help='invoices.json 文件路径')
    parser.add_argument('-o', '--output', help='输出Excel文件路径', default=None)
    parser.add_argument('-c', '--company', help='公司名称', default=None)
    parser.add_argument('-a', '--applicant', help='报销人姓名', default=None)

    args = parser.parse_args()

    # 生成输出文件名
    if args.output:
        output_file = args.output
    else:
        json_path = Path(args.json_file)
        today = datetime.now().strftime('%Y%m%d')
        output_file = json_path.parent / f"费用报销明细表_{today}.xlsx"

    create_expense_report(
        args.json_file,
        output_file,
        args.company,
        args.applicant
    )
